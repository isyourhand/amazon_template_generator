import openpyxl
import random
import traceback
import time
import os
import string
from tkinter import messagebox
from utilities.Calculate import fabric_add_prefix, has_numbers, get_Price, calculate_Price, replace_Url, current_Time, sale_End_Time
from utilities.Generate_Keyword import generate_Bullet, generate_Generic, generate_Title, customize_Generation
from utilities.random_Word.random_Word import style_name_random, neck_style_random, closure_type_random, sleeve_type_random, occasion_random, style_keywords_random, lining_description_random, theme_random, fit_type_random, top_style_random, lifecycle_supply_type_random, occasion_types_random, waist_style_random, item_weight_random, embellishment_feature1_random, bottom_style_random, leg_style_random, strap_type_random, fabric_wash_random, weave_type_random, collar_style_random, belt_style_random, material_composition_random, outer_material_type_random, item_length_description_random, inner_material_type_random, material_type_random, sport_type_random, pocket_description_random, control_type_random, front_style_random,rise_style_random
from utilities.excelDataProcess.excel_utils import load_and_validate_excel_data

from DataMap.DataMapping import us_size_map_Mapping, us_color_Map, column_map, seasons_map, sleeve_map, us_Size_Mapping, Product_Type_Map_New, department_name_Map, site_Map, age_group_map, gender_map, pattern_map, clothing_style_map, closure_map, column_headers

def copy_data_ca(target_file_path, source_file_path, prefix, 
              url_prefix, brand, site, 
              key_word_array1, key_word_array2, key_word_array3, key_word_array4,
              product_Type, seasons, department_name, age_group, gender, 
              patterns, nodes_array, 
              shipping_cost, title_include_fields, profit):
              
    try:
        # 打开源文件
        required_columns = ["SKU", "颜色", "尺寸", "水洗唛材质", "亚马逊建议价"] + [f"代理链接 {i}" for i in range(1, 10)]

        # 调用通用函数加载和验证数据
        source_wb, source_sheet, target_wb, target_sheet, data, column_indices = load_and_validate_excel_data(
            source_file_path=source_file_path,
            source_sheet_name='Sheet0 (2)',
            target_file_path=target_file_path,
            target_sheet_name='Sheet1',
            column_map=column_map,
            required_columns=required_columns,
            column_headers=column_headers
        )

        # 如果 source_wb 为 None，则表示整个加载和验证过程失败
        # 因为 load_and_validate_excel_data 在任何错误时返回全 None
        if source_wb is None:  # 如果返回 None，说明有错误，已由函数内部处理
            return

        # 绑定 sku
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "sku_bind"
        ws.cell(row=1, column=1, value="SKU")
        ws.cell(row=1, column=2, value="平台SKU")

        # 写入数据
        current_parent_sku = None
        model_value = None
        model_name_value = None
        part_number_value = None
        last_color = None
        last_sku_no_prefix = None
        
        color_prefix_map = {}
        color_prefix = "01-"

        for i, item in enumerate(data, start=4):
            sku = f"{prefix}{item['SKU']}"
            price_Hash = get_Price(item["亚马逊建议价"])
            price_with_margin = calculate_Price(price_Hash["加拿大"], shipping_cost, profit)
            temp_link = item["代理链接"][0]
            item["代理链接"][0] = item["代理链接"][1]
            item["代理链接"][1] = temp_link
            
            # IF
            item_type = "" if not nodes_array else random.choice(nodes_array)
            size = item['尺寸'] if item['尺寸'] not in us_Size_Mapping else us_Size_Mapping[item['尺寸']]
            size_map = item['尺寸'] if item['尺寸'] not in us_size_map_Mapping else us_size_map_Mapping[item['尺寸']]
            _color_ = item["颜色"] if item["颜色"].lower() not in us_color_Map else us_color_Map[item["颜色"].lower()]
            size_class = "Alpha" if not has_numbers(item['尺寸']) else "Age"
            height_type = "Regular" if not has_numbers(item['尺寸']) else ""

            # 检查后 5 位是否包含英文字符
            last_three = sku[-5:]
            contains_alpha = any(char.isalpha() for char in last_three)

            # Write Link
            target_sheet.cell(row=i, column=column_indices["main_image_url"], value=replace_Url(item["代理链接"][0], url_prefix))
            target_sheet.cell(row=i, column=column_indices["swatch_image_url"], value=replace_Url(item["代理链接"][0], url_prefix))
            for idx, link in enumerate(item["代理链接"]):
                if idx == 0 or item["代理链接"][idx] is None:
                    continue
                target_sheet.cell(row=i,column=column_indices[f"other_image_url{idx}"],value=replace_Url(item["代理链接"][idx], url_prefix))

            # New feature
            target_sheet.cell(row=i, column=column_indices["theme"], value=theme_random())
            target_sheet.cell(row=i, column=column_indices["style_keywords"], value=style_keywords_random())
            target_sheet.cell(row=i, column=column_indices["occasion"], value=occasion_random())
            target_sheet.cell(row=i, column=column_indices["fulfillment_latency"], value=2)

            # Other
            target_sheet.cell(row=i, column=column_indices["feed_product_type"], value=Product_Type_Map_New[product_Type])
            target_sheet.cell(row=i, column=column_indices["variation_theme"], value="colorsize")
            target_sheet.cell(row=i, column=column_indices["department_name"], value=department_name_Map[department_name])
            target_sheet.cell(row=i, column=column_indices["age_range_description"], value=age_group_map[age_group])
            target_sheet.cell(row=i, column=column_indices["target_gender"], value=gender_map[gender])
            target_sheet.cell(row=i, column=column_indices["import_designation"], value="Imported")
            target_sheet.cell(row=i, column=column_indices["care_instructions"], value="Machine Wash or Hand Wash")
            target_sheet.cell(row=i, column=column_indices["country_of_origin"], value="China")
            target_sheet.cell(row=i, column=column_indices["item_type"], value=item_type)
            target_sheet.cell(row=i, column=column_indices["closure_type"], value=closure_type_random())
            target_sheet.cell(row=i, column=column_indices["sport_type1"], value=sport_type_random())
            target_sheet.cell(row=i, column=column_indices["currency"], value="CAD")

            target_sheet.cell(row=i, column=column_indices["item_type_name"], value=f"{department_name_Map[department_name]} {Product_Type_Map_New[product_Type]}")
            target_sheet.cell(row=i, column=column_indices["band_size_num"], value="1")
            target_sheet.cell(row=i, column=column_indices["band_size_num_unit_of_measure"], value="centimeters")
            target_sheet.cell(row=i, column=column_indices["supplier_declared_dg_hz_regulation1"], value="Not Applicable")
            target_sheet.cell(row=i, column=column_indices["supplier_declared_material_regulation1"], value="Not Applicable")
            target_sheet.cell(row=i, column=column_indices["country_as_labeled"], value="CN")
            target_sheet.cell(row=i, column=column_indices["number_of_pieces"], value="1")
            target_sheet.cell(row=i, column=column_indices["water_resistance_level"], value="water_resistant")
            target_sheet.cell(row=i, column=column_indices["target_audience_keywords1"], value=department_name_Map[department_name])
            target_sheet.cell(row=i, column=column_indices["target_audience_keywords2"], value=department_name_Map[department_name])
            target_sheet.cell(row=i, column=column_indices["target_audience_keywords3"], value=age_group_map[age_group])
            target_sheet.cell(row=i, column=column_indices["target_audience_keywords4"], value=gender_map[gender])
            target_sheet.cell(row=i, column=column_indices["target_audience_keywords5"], value=department_name_Map[department_name])
            target_sheet.cell(row=i, column=column_indices["is_autographed"], value="no")
            
            # 动态写入单元格
            for idx in range(1, 6):
                column_key = f"occasion_type{idx}"
                if column_key in column_indices:  # 确保列索引存在
                    target_sheet.cell(row=i, column=column_indices[column_key], value=occasion_types_random())
            
            for idx in range(1, 6):
                column_key = f"fit_type{idx}"
                if column_key in column_indices:  # 确保列索引存在
                    target_sheet.cell(row=i, column=column_indices[column_key], value=fit_type_random())

            for idx in range(1, 5):
                column_key = f"platinum_keywords{idx}"
                if column_key in column_indices:  # 确保列索引存在
                    target_sheet.cell(row=i, column=column_indices[column_key], value=customize_Generation(key_word_array1, key_word_array2, 49))

            # 定义所有的 lifestyle 类型
            lifestyle_types = [
                "Maternity", "Kids Playwear", "Nursing", "Dress", "Work Utility", 
                "Athletic", "Casual", "Uniform", "Themed", "Exotic"
            ]

            # 动态写入 lifestyle 类型
            for idx, lifestyle in enumerate(lifestyle_types, start=1):
                column_key = f"lifestyle{idx}"
                if column_key in column_indices:  # 确保列索引存在
                    target_sheet.cell(row=i, column=column_indices[column_key], value=lifestyle)

            # 定义所有的 special_features 类型
            special_features = [
                "Breathable", "Lightweight", "Moisture Wicking", "Quick Dry", "Sun Protection", 
                "Adjustable", "Convertible", "Removable Padding", "Reversible", "Stretchable", 
                "Sustainable", "Wrinkle Free"
            ]

            # 动态写入 special_features 类型
            for idx, feature in enumerate(special_features, start=1):
                column_key = f"special_features{idx}"
                if column_key in column_indices:  # 确保列索引存在
                    target_sheet.cell(row=i, column=column_indices[column_key], value=feature)

            target_sheet.cell(row=i, column=column_indices["sport_type2"], value=sport_type_random())
            target_sheet.cell(row=i, column=column_indices["seasons1"], value="Spring")
            target_sheet.cell(row=i, column=column_indices["seasons2"], value="Summer")
            target_sheet.cell(row=i, column=column_indices["seasons3"], value="Fall")
            target_sheet.cell(row=i, column=column_indices["seasons4"], value="Winter")
            target_sheet.cell(row=i, column=column_indices["collection_name"], value=seasons_map[random.choice(seasons)])
            target_sheet.cell(row=i, column=column_indices["weave_type"], value=weave_type_random())
            target_sheet.cell(row=i, column=column_indices["lifecycle_supply_type"], value=lifecycle_supply_type_random())
            target_sheet.cell(row=i, column=column_indices["collar_style"], value=collar_style_random())
            target_sheet.cell(row=i, column=column_indices["number_of_boxes"], value="1")
            target_sheet.cell(row=i, column=column_indices["embellishment_feature1"], value=embellishment_feature1_random())
            target_sheet.cell(row=i, column=column_indices["bottom_style"], value=bottom_style_random())
            target_sheet.cell(row=i, column=column_indices["leg_style"], value=leg_style_random())
            target_sheet.cell(row=i, column=column_indices["strap_type"], value=strap_type_random())
            target_sheet.cell(row=i, column=column_indices["fabric_wash"], value=fabric_wash_random())
            target_sheet.cell(row=i, column=column_indices["number_of_pockets"], value="2")
            target_sheet.cell(row=i, column=column_indices["waist_style"], value=waist_style_random())
            target_sheet.cell(row=i, column=column_indices["cpsia_cautionary_statement1"], value="no_warning_applicable")
            target_sheet.cell(row=i, column=column_indices["item_weight_unit_of_measure"], value="GR")
            target_sheet.cell(row=i, column=column_indices["item_weight"], value=item_weight_random())

            # Cloth detail
            target_sheet.cell(row=i, column=column_indices["neck_style"], value=neck_style_random())
            target_sheet.cell(row=i, column=column_indices["pattern_type"], value=pattern_map[random.choice(patterns)])
            target_sheet.cell(row=i, column=column_indices["style_name"], value=style_name_random())
            target_sheet.cell(row=i, column=column_indices["outer_material_type1"], value=outer_material_type_random())
            target_sheet.cell(row=i, column=column_indices["material_composition"], value=material_composition_random())
            target_sheet.cell(row=i, column=column_indices["item_length_description"], value=item_length_description_random())
            target_sheet.cell(row=i, column=column_indices["outer_material_type2"], value=outer_material_type_random())
            target_sheet.cell(row=i, column=column_indices["outer_material_type3"], value=outer_material_type_random())
            target_sheet.cell(row=i, column=column_indices["outer_material_type4"], value=outer_material_type_random())
            target_sheet.cell(row=i, column=column_indices["outer_material_type5"], value=outer_material_type_random())
            target_sheet.cell(row=i, column=column_indices["inner_material_type"], value=inner_material_type_random())
            target_sheet.cell(row=i, column=column_indices["material_type1"], value="Cotton")
            target_sheet.cell(row=i, column=column_indices["fabric_type"], value=fabric_add_prefix(item["水洗唛材质"]))
            target_sheet.cell(row=i, column=column_indices["fur_description"], value=fabric_add_prefix(item["水洗唛材质"]))
            target_sheet.cell(row=i, column=column_indices["pocket_description"], value=pocket_description_random())
            target_sheet.cell(row=i, column=column_indices["pattern_name"], value=pattern_map[random.choice(patterns)])
            target_sheet.cell(row=i, column=column_indices["sleeve_type"], value=sleeve_type_random())
            target_sheet.cell(row=i, column=column_indices["top_style"], value=top_style_random())
            target_sheet.cell(row=i, column=column_indices["belt_style"], value=belt_style_random())
            target_sheet.cell(row=i, column=column_indices["material_type2"], value="Cashmere")
            target_sheet.cell(row=i, column=column_indices["material_type3"], value=material_type_random())
            target_sheet.cell(row=i, column=column_indices["control_type"], value=control_type_random())
            target_sheet.cell(row=i, column=column_indices["front_style"], value=front_style_random())
            target_sheet.cell(row=i, column=column_indices["rise_style"], value=rise_style_random())
            target_sheet.cell(row=i, column=column_indices["lining_description"], value=lining_description_random())

            # Sell_related
            target_sheet.cell(row=i, column=column_indices["quantity"], value=199)
            target_sheet.cell(row=i, column=column_indices["merchant_shipping_group_name"], value=shipping_cost)
            target_sheet.cell(row=i, column=column_indices["item_package_quantity"], value=1)
            target_sheet.cell(row=i, column=column_indices["number_of_items"], value=1)
            target_sheet.cell(row=i, column=column_indices["condition_type"], value="New")

            # Write Brand
            target_sheet.cell(row=i, column=column_indices["brand_name"], value=brand)
            target_sheet.cell(row=i, column=column_indices["manufacturer"], value=brand + Product_Type_Map_New[product_Type] + "Factory")
            

            # Write Key Word
            target_sheet.cell(row=i, column=column_indices["bullet_point1"], value=generate_Bullet(key_word_array3,key_word_array4))
            target_sheet.cell(row=i, column=column_indices["bullet_point2"], value=generate_Bullet(key_word_array3,key_word_array4))
            target_sheet.cell(row=i, column=column_indices["bullet_point3"], value=generate_Bullet(key_word_array3,key_word_array4))
            target_sheet.cell(row=i, column=column_indices["bullet_point4"], value=generate_Bullet(key_word_array3,key_word_array4))
            target_sheet.cell(row=i, column=column_indices["bullet_point5"], value=generate_Bullet(key_word_array3,key_word_array4))
            target_sheet.cell(row=i, column=column_indices["product_description"], value=generate_Bullet(key_word_array3,key_word_array4))
            target_sheet.cell(row=i, column=column_indices["generic_keywords"], value=generate_Generic(key_word_array3,key_word_array4))

            target_sheet.cell(row=i, column=column_indices["item_name"], value=generate_Title(key_word_array1, key_word_array2, Product_Type_Map_New[product_Type], pattern_map[random.choice(patterns)], _color_, size_map, item["水洗唛材质"], title_include_fields))

            ws.cell(row=i-2, column=1, value=item['SKU'])
            ws.cell(row=i-2, column=2, value=sku)

            

            # When Child
            if contains_alpha:  # Child
                
                # subTitle
                


                # Sku_Ralated
                target_sheet.cell(row=i, column=column_indices["item_sku"], value=sku)
                target_sheet.cell(row=i, column=column_indices["parent_child"], value="Child")
                target_sheet.cell(row=i, column=column_indices["parent_sku"], value=current_parent_sku)
                target_sheet.cell(row=i, column=column_indices["relationship_type"], value="Variation")

                # Price_Related
                target_sheet.cell(row=i, column=column_indices["standard_price"], value=price_with_margin)
                target_sheet.cell(row=i, column=column_indices["product_site_launch_date"], value=current_Time())
                target_sheet.cell(row=i, column=column_indices["merchant_release_date"], value=current_Time())
                target_sheet.cell(row=i, column=column_indices["sale_from_date"], value=current_Time())
                target_sheet.cell(row=i, column=column_indices["sale_end_date"], value=sale_End_Time())
                target_sheet.cell(row=i, column=column_indices["list_price"], value=price_with_margin+10)
                target_sheet.cell(row=i, column=column_indices["sale_price"], value=price_with_margin-1)

                # Size_Related
                target_sheet.cell(row=i, column=column_indices["size_name"], value=size_map)
                target_sheet.cell(row=i, column=column_indices["size_map"], value=size_map)
                target_sheet.cell(row=i, column=column_indices["apparel_size"], value=size)
                target_sheet.cell(row=i, column=column_indices["shirt_size"], value=size)
                target_sheet.cell(row=i, column=column_indices["bottoms_size"], value=size)
                target_sheet.cell(row=i, column=column_indices["skirt_size"], value=size)
                target_sheet.cell(row=i, column=column_indices["apparel_body_type"], value="Regular")
                target_sheet.cell(row=i, column=column_indices["apparel_height_type"], value=height_type)
                target_sheet.cell(row=i, column=column_indices["apparel_size_system"], value=site_Map[site])
                target_sheet.cell(row=i, column=column_indices["apparel_size_class"], value=size_class)
                target_sheet.cell(row=i, column=column_indices["shirt_size_class"], value=size_class)
                target_sheet.cell(row=i, column=column_indices["shirt_size_system"], value=site_Map[site])
                target_sheet.cell(row=i, column=column_indices["shirt_body_type"], value="Regular")
                target_sheet.cell(row=i, column=column_indices["shirt_height_type"], value=height_type)
                target_sheet.cell(row=i, column=column_indices["bottoms_size_system"], value=site_Map[site])
                target_sheet.cell(row=i, column=column_indices["bottoms_size_class"], value=size_class)
                target_sheet.cell(row=i, column=column_indices["bottoms_body_type"], value="Regular")
                target_sheet.cell(row=i, column=column_indices["bottoms_height_type"], value=height_type)
                target_sheet.cell(row=i, column=column_indices["skirt_size_system"], value="CA")
                target_sheet.cell(row=i, column=column_indices["skirt_size_class"], value=size_class)
                target_sheet.cell(row=i, column=column_indices["skirt_body_type"], value="Regular")
                target_sheet.cell(row=i, column=column_indices["skirt_height_type"], value=height_type)
                target_sheet.cell(row=i, column=column_indices["shapewear_size_system"], value="CA")
                target_sheet.cell(row=i, column=column_indices["shapewear_size_class"], value=size_class)
                target_sheet.cell(row=i, column=column_indices["shapewear_size"], value=size)
                target_sheet.cell(row=i, column=column_indices["shapewear_body_type"], value="Regular")
                target_sheet.cell(row=i, column=column_indices["shapewear_height_type"], value=height_type)
                
                # 计算当前无前缀 sku 长度
                import re
                pattern = r'([A-Za-z]+)(\d+)'
                matches = re.match(pattern, item['SKU'])
                no_prefix_sku_length = len(matches.group(1) + matches.group(2)) + 1


                color = item["颜色"]
                if color:
                    if item['SKU'][:no_prefix_sku_length] != last_sku_no_prefix or color != last_color:
                        color_prefix_map[color] = color_prefix_map.get(color, 0) + 1
                        color_prefix = f"{color_prefix_map[color]:02d}-"
                        last_color = color
                        last_sku_no_prefix = item['SKU'][:no_prefix_sku_length]

                    target_sheet.cell(row=i, column=column_indices["color_map"], value=_color_)
                    target_sheet.cell(row=i, column=column_indices["color_name"], value=color_prefix + color)
                
                target_sheet.cell(row=i, column=column_indices["model"], value=model_value)
                target_sheet.cell(row=i, column=column_indices["model_name"], value=model_name_value)
                target_sheet.cell(row=i, column=column_indices["part_number"], value=part_number_value)
            else:  # Parent
                current_parent_sku = sku

                random_number = ''.join(random.choices(string.digits, k=4))
                model_value = customize_Generation(key_word_array1, key_word_array2, 33) + random_number
                model_name_value = customize_Generation(key_word_array1, key_word_array2, 105)
                part_number_value = customize_Generation(key_word_array1, key_word_array2, 32) + random_number

                target_sheet.cell(row=i, column=column_indices["model"], value=model_value)
                target_sheet.cell(row=i, column=column_indices["model_name"], value=model_name_value)
                target_sheet.cell(row=i, column=column_indices["part_number"], value=part_number_value)

                target_sheet.cell(row=i, column=column_indices["item_sku"], value=sku)
                target_sheet.cell(row=i, column=column_indices["parent_child"], value="Parent")

                last_color = None
                last_sku_no_prefix = None
                color_prefix_map = {}

        target_wb.save(target_file_path)
        output_file = os.path.join(os.path.dirname(target_file_path), f"{prefix}_sku_binding.xlsx")
        wb.save(output_file)
        messagebox.showinfo("成功", f"数据处理完成！\n\n目标文件: \n{target_file_path}\n\nSKU绑定文件: \n{output_file}")

    except PermissionError:
        messagebox.showerror("权限错误", 
                           f"无法保存文件到:\n{target_file_path}\n\n请检查文件是否被占用或打开")
    except Exception as e:
        error_type = type(e).__name__
        error_msg = str(e)
        user_message = "发生未知错误，请联系技术支持并提供以下信息：\n\n" \
                      f"错误类型: {error_type}\n" \
                      f"错误详情: {error_msg}\n" \
                      "请检查输入文件格式并重试操作"
        messagebox.showerror("处理错误", user_message)
        print(f"处理文件时发生错误:\n"
              f"错误类型: {error_type}\n"
              f"错误信息: {error_msg}\n"
              f"详细堆栈跟踪:\n{traceback.format_exc()}")