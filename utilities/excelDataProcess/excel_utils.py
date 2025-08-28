# excel_utils.py
import openpyxl
from tkinter import messagebox
from typing import Dict, List, Tuple, Optional

def load_workbook_safe(file_path: str, sheet_name: str, error_title: str) -> Tuple[Optional[openpyxl.Workbook], Optional[openpyxl.worksheet.worksheet.Worksheet]]:
    """安全加载 Excel 工作簿并返回指定工作表。
    
    Args:
        file_path: Excel 文件路径
        sheet_name: 要加载的工作表名称
        error_title: 错误消息的标题
        
    Returns:
        包含工作簿和工作表的元组，如果加载失败则返回 (None, None)
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook[sheet_name]
        return workbook, worksheet
    except FileNotFoundError:
        messagebox.showerror(f"{error_title} - 文件错误",
                           f"找不到文件: {file_path}\n请检查文件路径是否正确。")
        return None, None
    except KeyError:
        messagebox.showerror(f"{error_title} - 工作表错误",
                           f"{error_title}中找不到工作表 '{sheet_name}'\n请确认工作表名称是否正确。")
        return None, None

def validate_worksheet_content(worksheet: openpyxl.worksheet.worksheet.Worksheet, file_path: str, sheet_name: str, file_type: str) -> bool:
    """验证工作表是否包含数据。"""
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        messagebox.showerror(f"{file_type}错误",
                           f"'{file_path}' 中的工作表 '{sheet_name}' 为空。\n请检查文件内容。")
        return False
    return True

def get_column_mappings(worksheet: openpyxl.worksheet.worksheet.Worksheet, column_map: Dict) -> Dict:
    """将列标题映射到对应的列号。"""
    for cell in worksheet[1]:
        if cell.value in column_map:
            column_map[cell.value] = cell.column
    return column_map

def extract_row_data(row: Tuple, column_map: Dict) -> Dict:
    """从单行提取并结构化数据。"""
    agent_links = [row[column_map[f"代理链接 {i}"] - 1].value for i in range(1, 10)]
    return {
        "SKU": row[column_map["SKU"] - 1].value,
        "颜色": row[column_map["颜色"] - 1].value,
        "尺寸": row[column_map["尺寸"] - 1].value,
        "水洗唛材质": row[column_map["水洗唛材质"] - 1].value,
        "亚马逊建议价": row[column_map["亚马逊建议价"] - 1].value,
        "代理链接": agent_links
    }

def validate_required_fields(row_data: Dict, row_num: int) -> Optional[List[str]]:
    """检查行数据中是否存在缺失的必要字段。"""
    required_fields = {
        "SKU": row_data["SKU"],
        "颜色": row_data["颜色"],
        "尺寸": row_data["尺寸"],
        "水洗唛材质": row_data["水洗唛材质"],
        "亚马逊建议价": row_data["亚马逊建议价"],
        "代理链接 2": row_data["代理链接"][1]
    }
    empty_fields = [field for field, value in required_fields.items()
                   if value is None or (isinstance(value, str) and value.strip() == "")]
    if empty_fields:
        messagebox.showerror("数据错误",
                           f"源文件第 {row_num} 行\n({', '.join(empty_fields)})\n缺少必要字段\n请检查数据的完整性。")
        return empty_fields
    return None

def load_and_validate_excel_data(
    source_file_path: str, source_sheet_name: str,
    target_file_path: str, target_sheet_name: str,
    column_map: Dict, required_columns: List, column_headers: List
) -> Tuple[Optional[openpyxl.Workbook], Optional[openpyxl.worksheet.worksheet.Worksheet],
          Optional[openpyxl.Workbook], Optional[openpyxl.worksheet.worksheet.Worksheet],
          Optional[List[Dict]], Optional[Dict]]:
    """
    加载并验证源文件和目标文件的 Excel 数据。
    
    Args:
        source_file_path: 源 Excel 文件路径
        source_sheet_name: 源工作表名称
        target_file_path: 目标 Excel 文件路径
        target_sheet_name: 目标工作表名称
        column_map: 列名称到索引的映射字典
        required_columns: 必要的列名称列表
        column_headers: 目标文件中预期的列标题列表
        
    Returns:
        元组 (源工作簿, 源工作表, 目标工作簿, 目标工作表, 数据, 列索引)
    """
    # 加载源文件
    source_wb, source_sheet = load_workbook_safe(source_file_path, source_sheet_name, "源文件")
    if not source_wb or not source_sheet or not validate_worksheet_content(source_sheet, source_file_path, source_sheet_name, "源文件"):
        return None, None, None, None, None, None

    # 加载目标文件
    target_wb, target_sheet = load_workbook_safe(target_file_path, target_sheet_name, "目标文件")
    if not target_wb or not target_sheet or not validate_worksheet_content(target_sheet, target_file_path, target_sheet_name, "目标文件"):
        return None, None, None, None, None, None

    # 映射源文件列
    column_map = get_column_mappings(source_sheet, column_map)
    missing_columns = [col for col, idx in column_map.items() if idx is None and col in required_columns]
    if missing_columns:
        messagebox.showerror("列缺失错误",
                           f"源文件中缺少必要列：\n{', '.join(missing_columns)}\n请检查源文件格式。")
        return None, None, None, None, None, None

    # 提取源数据
    data = []
    for row in source_sheet.iter_rows(min_row=2, max_row=source_sheet.max_row):
        if row[column_map["SKU"] - 1].value:
            try:
                row_data = extract_row_data(row, column_map)
                if validate_required_fields(row_data, row[0].row):
                    return None, None, None, None, None, None
                data.append(row_data)
            except IndexError as e:
                messagebox.showerror("数据错误",
                                   f"第 {row[0].row} 行：数据不完整，缺少某些列的值。\n错误详情: {str(e)}")
                return None, None, None, None, None, None

    if not data:
        messagebox.showerror("数据错误",
                           f"源文件 '{source_file_path}' 中没有有效数据。\n请检查文件内容。")
        return None, None, None, None, None, None

    # 验证目标文件结构
    if target_sheet.max_row < 3:
        messagebox.showerror("目标文件错误",
                           f"目标文件 '{target_file_path}' 的工作表 '{target_sheet_name}' 数据不足。\n请检查文件内容。")
        return None, None, None, None, None, None

    # 映射目标文件列
    column_indices = {header: None for header in column_headers}
    for cell in target_sheet[3]:
        if cell.value in column_indices:
            column_indices[cell.value] = cell.column

    missing_headers = [header for header, idx in column_indices.items() if idx is None]
    if missing_headers:
        messagebox.showerror("列缺失错误",
                           f"目标文件中缺少必要列：\n{', '.join(missing_headers)}\n请确认目标文件是否对应所选站点。")
        return None, None, None, None, None, None

    return source_wb, source_sheet, target_wb, target_sheet, data, column_indices