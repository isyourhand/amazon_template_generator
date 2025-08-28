import openpyxl
import random
import time
import os
from tkinter import messagebox

import openpyxl.workbook
from utilities.Calculate import get_Price, calculate_Price, replace_Url, current_Time, sale_End_Time
from utilities.Generate_Keyword import generate_Bullet, generate_Generic, generate_Title
from utilities.random_Word import waist_style_random, cup_size_random, item_weight_random, embellishment_feature1_random, product_benefit_random, bottom_style_random, leg_style_random, strap_type_random, underwire_type_random, fabric_wash_random, weave_type_random, sport_type2_random, special_features_random, belt_style_random, material_compositon_random, outer_material_type1_random, lifestyle1_random, item_length_description_random, inner_material_type1_random, material_type1_random, sport_type1_random, pocket_description_random, pattern_name_random, material_type2_random, material_type3_random, control_type_random, front_style_random,rise_style_random,toe_style_random

from DataMap.DataMapping import us_color_Map, waist_style_map, leg_style_map, column_map, seasons_map, us_Size_Mapping, Product_Type_Map, target_Audience_Map, site_Map, age_group_map, gender_map, pattern_map, skirt_length_map, clothing_style_map, fit_map, closure_map, column_headers, pocket_map

def copy_data(target_file_path, source_file_path, prefix, 
              url_prefix, brand, site, 
              key_word_array1, key_word_array2, key_word_array3, key_word_array4,
              product_Type, seasons, target_audience, age_group, gender, 
              patterns, styles, skirt_length, fit_types, nodes_array, 
              closure_type, pocket_type, leg_styles, waist_style, shipping_cost, title_include_fields):
              
    try:
        # 打开源文件
        source_wb = openpyxl.load_workbook(source_file_path)
        source_sheet = source_wb['Sheet0 (2)']

        # 打开目标文件
        target_wb = openpyxl.load_workbook(target_file_path)
        target_sheet = target_wb['Sheet1']

        # 绑定 sku
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "sku_bind"
        ws.cell(row=1, column=1, value="SKU")
        ws.cell(row=1, column=2, value="平台SKU")

        # 获取源文件相关列数据
        
        for cell in source_sheet[1]:
            if cell.value in column_map:
                column_map[cell.value] = cell.column

        if None in column_map.values():
            missing_columns = [key for key, value in column_map.items() if value is None]
            messagebox.showerror("错误", f"源文件中缺失列: {', '.join(missing_columns)}")
            return

        # 提取源文件数据
        data = []
        for row in source_sheet.iter_rows(min_row=2, max_row=source_sheet.max_row):
            if row[column_map["SKU"] - 1].value:
                data.append({
                    "SKU": row[column_map["SKU"] - 1].value,
                    "颜色": row[column_map["颜色"] - 1].value,
                    "尺寸": row[column_map["尺寸"] - 1].value,
                    "水洗唛材质": row[column_map["水洗唛材质"] - 1].value,
                    "亚马逊建议价": row[column_map["亚马逊建议价"] - 1].value,
                    "代理链接": [
                        row[column_map[f"代理链接 {i}"] - 1].value for i in range(1, 10)
                    ]
                })


        # 获取目标文件的列
        
        column_indices = {header: None for header in column_headers}

        for cell in target_sheet[3]:
            if cell.value in column_indices:
                column_indices[cell.value] = cell.column

        if None in column_indices.values():
            missing_headers = [key for key, value in column_indices.items() if value is None]
            messagebox.showerror("错误", f"目标文件缺少列: {', '.join(missing_headers)}")
            return

        # 写入数据
        current_parent_sku = None
        last_color = None
        last_sku_no_prefix = None

        color_prefix_map = {}
        color_prefix = "01-"

        for i, item in enumerate(data, start=4):
            sku = f"{prefix}{item['SKU']}"
            price_Hash = get_Price(item["亚马逊建议价"])
            price_with_margin = calculate_Price(price_Hash["美国"], shipping_cost)
            temp_link = item["代理链接"][0]
            item["代理链接"][0] = item["代理链接"][1]
            item["代理链接"][1] = temp_link
            
            # IF
            skirt_length_real = skirt_length_map[skirt_length] if skirt_length != "并非裙子" else item_length_description_random()
            pocket_description = pocket_map[pocket_type] if pocket_type != "无" else pocket_description_random()
            item_type = "" if not nodes_array else random.choice(nodes_array)
            _waist_style_ = waist_style_map[waist_style] if waist_style != "皆可" else waist_style_random()
            size = item['尺寸'] if item['尺寸'] not in us_Size_Mapping else us_Size_Mapping[item['尺寸']]
            _color_ = item["颜色"] if item["颜色"].lower() not in us_color_Map else us_color_Map[item["颜色"].lower()]

            # 检查后三位是否包含英文字符
            last_three = sku[-3:]
            contains_alpha = any(char.isalpha() for char in last_three)

            # Write Link
            target_sheet.cell(row=i, column=column_indices["main_image_url"], value=replace_Url(item["代理链接"][0], url_prefix))
            target_sheet.cell(row=i, column=column_indices["swatch_image_url"], value=replace_Url(item["代理链接"][0], url_prefix))
            for idx, link in enumerate(item["代理链接"]):
                if idx == 0 or item["代理链接"][idx] is None:
                    continue
                target_sheet.cell(row=i,column=column_indices[f"other_image_url{idx}"],value=replace_Url(item["代理链接"][idx], url_prefix))

            # Other
            target_sheet.cell(row=i, column=column_indices["feed_product_type"], value=Product_Type_Map[product_Type])
            target_sheet.cell(row=i, column=column_indices["variation_theme"], value="colorsize")
            target_sheet.cell(row=i, column=column_indices["department_name"], value=target_Audience_Map[target_audience])
            target_sheet.cell(row=i, column=column_indices["age_range_description"], value=age_group_map[age_group])
            target_sheet.cell(row=i, column=column_indices["target_gender"], value=gender_map[gender])
            target_sheet.cell(row=i, column=column_indices["is_adult_product"], value="No")
            target_sheet.cell(row=i, column=column_indices["lifestyle1"], value=lifestyle1_random())
            target_sheet.cell(row=i, column=column_indices["import_designation"], value="Imported")
            target_sheet.cell(row=i, column=column_indices["care_instructions"], value="Machine Wash or Hand Wash")
            target_sheet.cell(row=i, column=column_indices["country_of_origin"], value="China")
            target_sheet.cell(row=i, column=column_indices["fit_type1"], value=fit_map[random.choice(fit_types)])
            target_sheet.cell(row=i, column=column_indices["item_type"], value=item_type)
            target_sheet.cell(row=i, column=column_indices["closure_type"], value=closure_map[closure_type])
            target_sheet.cell(row=i, column=column_indices["sport_type1"], value=sport_type1_random())
            target_sheet.cell(row=i, column=column_indices["currency"], value="USD")
            target_sheet.cell(row=i, column=column_indices["league_name"], value="World Baseball Classic")
            target_sheet.cell(row=i, column=column_indices["special_features1"], value=special_features_random())
            target_sheet.cell(row=i, column=column_indices["special_features2"], value=special_features_random())
            target_sheet.cell(row=i, column=column_indices["special_features3"], value=special_features_random())
            target_sheet.cell(row=i, column=column_indices["special_features4"], value=special_features_random())
            target_sheet.cell(row=i, column=column_indices["special_features5"], value=special_features_random())
            target_sheet.cell(row=i, column=column_indices["item_type_name"], value=f"{target_Audience_Map[target_audience]} {Product_Type_Map[product_Type]}")
            target_sheet.cell(row=i, column=column_indices["band_size_num"], value="1")
            target_sheet.cell(row=i, column=column_indices["band_size_num_unit_of_measure"], value="centimeters")
            target_sheet.cell(row=i, column=column_indices["batteries_required"], value="no")
            target_sheet.cell(row=i, column=column_indices["supplier_declared_dg_hz_regulation1"], value="Not Applicable")
            target_sheet.cell(row=i, column=column_indices["supplier_declared_material_regulation1"], value="Not Applicable")
            target_sheet.cell(row=i, column=column_indices["country_as_labeled"], value="CN")
            target_sheet.cell(row=i, column=column_indices["number_of_pieces"], value="1")
            target_sheet.cell(row=i, column=column_indices["water_resistance_level"], value="Waterproof")
            target_sheet.cell(row=i, column=column_indices["target_audience_keywords1"], value=target_Audience_Map[target_audience])
            target_sheet.cell(row=i, column=column_indices["target_audience_keywords2"], value=target_Audience_Map[target_audience])
            target_sheet.cell(row=i, column=column_indices["target_audience_keywords3"], value=age_group_map[age_group])
            target_sheet.cell(row=i, column=column_indices["target_audience_keywords4"], value=gender_map[gender])
            target_sheet.cell(row=i, column=column_indices["target_audience_keywords5"], value=target_Audience_Map[target_audience])
            target_sheet.cell(row=i, column=column_indices["is_autographed"], value="no")
            target_sheet.cell(row=i, column=column_indices["occasion_type1"], value="Christmas")
            target_sheet.cell(row=i, column=column_indices["occasion_type2"], value="Wedding")
            target_sheet.cell(row=i, column=column_indices["occasion_type3"], value="Cocktail")
            target_sheet.cell(row=i, column=column_indices["occasion_type4"], value="Birthday")
            target_sheet.cell(row=i, column=column_indices["occasion_type5"], value="Halloween")
            target_sheet.cell(row=i, column=column_indices["occasion_type6"], value="Baptism")
            target_sheet.cell(row=i, column=column_indices["occasion_type7"], value="Easter")
            target_sheet.cell(row=i, column=column_indices["occasion_type8"], value="Graduation")
            target_sheet.cell(row=i, column=column_indices["occasion_type9"], value="St.Patricks Day")
            target_sheet.cell(row=i, column=column_indices["occasion_type10"], value="Thanksgiving")
            target_sheet.cell(row=i, column=column_indices["occasion_type11"], value="Hanukkah")
            target_sheet.cell(row=i, column=column_indices["occasion_type12"], value="Memorial Day")
            target_sheet.cell(row=i, column=column_indices["occasion_type13"], value="Kwanzaa")
            target_sheet.cell(row=i, column=column_indices["occasion_type14"], value="Bachelorette")
            target_sheet.cell(row=i, column=column_indices["occasion_type15"], value="Bridal Shower")
            target_sheet.cell(row=i, column=column_indices["occasion_type16"], value="Engagement")
            target_sheet.cell(row=i, column=column_indices["occasion_type17"], value="Mothers Day")
            target_sheet.cell(row=i, column=column_indices["occasion_type18"], value="New Year")
            target_sheet.cell(row=i, column=column_indices["occasion_type19"], value="Independence Day")
            target_sheet.cell(row=i, column=column_indices["occasion_type20"], value="Valentines Day")
            target_sheet.cell(row=i, column=column_indices["occasion_type21"], value="Prom")
            target_sheet.cell(row=i, column=column_indices["occasion_type22"], value="Anniversary")
            target_sheet.cell(row=i, column=column_indices["occasion_type23"], value="Housewarming")
            target_sheet.cell(row=i, column=column_indices["occasion_type24"], value="Festival")
            target_sheet.cell(row=i, column=column_indices["occasion_type25"], value="Party")
            target_sheet.cell(row=i, column=column_indices["occasion_type26"], value="Mardi Gras")
            target_sheet.cell(row=i, column=column_indices["occasion_type27"], value="Reception")
            target_sheet.cell(row=i, column=column_indices["sport_type2"], value=sport_type2_random())
            target_sheet.cell(row=i, column=column_indices["seasons1"], value="Spring")
            target_sheet.cell(row=i, column=column_indices["seasons2"], value="Summer")
            target_sheet.cell(row=i, column=column_indices["seasons3"], value="Fall")
            target_sheet.cell(row=i, column=column_indices["seasons4"], value="Winter")
            target_sheet.cell(row=i, column=column_indices["collection_name"], value=seasons_map[random.choice(seasons)])
            target_sheet.cell(row=i, column=column_indices["weave_type"], value=weave_type_random())
            target_sheet.cell(row=i, column=column_indices["lifecycle_supply_type1"], value="Perennial")
            target_sheet.cell(row=i, column=column_indices["lifecycle_supply_type2"], value="Year Round Replenishable")
            target_sheet.cell(row=i, column=column_indices["lifecycle_supply_type3"], value="Fashion")
            target_sheet.cell(row=i, column=column_indices["number_of_boxes"], value="1")
            target_sheet.cell(row=i, column=column_indices["embellishment_feature1"], value=embellishment_feature1_random())
            target_sheet.cell(row=i, column=column_indices["product_benefit"], value=product_benefit_random())
            target_sheet.cell(row=i, column=column_indices["bottom_style"], value=bottom_style_random())
            target_sheet.cell(row=i, column=column_indices["leg_style"], value=leg_style_map[random.choice(leg_styles)])
            target_sheet.cell(row=i, column=column_indices["strap_type"], value=strap_type_random())
            target_sheet.cell(row=i, column=column_indices["underwire_type"], value=underwire_type_random())
            target_sheet.cell(row=i, column=column_indices["fabric_wash"], value=fabric_wash_random())
            target_sheet.cell(row=i, column=column_indices["number_of_pockets"], value="2")
            target_sheet.cell(row=i, column=column_indices["waist_style"], value=_waist_style_)
            target_sheet.cell(row=i, column=column_indices["cup_size"], value=cup_size_random())
            target_sheet.cell(row=i, column=column_indices["cpsia_cautionary_statement1"], value="NoWarningApplicable")
            target_sheet.cell(row=i, column=column_indices["item_weight_unit_of_measure"], value="GR")
            target_sheet.cell(row=i, column=column_indices["item_weight"], value=item_weight_random())

            # Cloth detail
            target_sheet.cell(row=i, column=column_indices["pattern_type"], value=pattern_map[random.choice(patterns)])
            target_sheet.cell(row=i, column=column_indices["style_name"], value=clothing_style_map[random.choice(styles)])
            target_sheet.cell(row=i, column=column_indices["outer_material_type1"], value=outer_material_type1_random())
            target_sheet.cell(row=i, column=column_indices["material_composition1"], value=material_compositon_random())
            target_sheet.cell(row=i, column=column_indices["item_length_description"], value=skirt_length_real)
            target_sheet.cell(row=i, column=column_indices["outer_material_type2"], value="Cotton")
            target_sheet.cell(row=i, column=column_indices["outer_material_type3"], value="Faux Leather")
            target_sheet.cell(row=i, column=column_indices["outer_material_type4"], value="Wool")
            target_sheet.cell(row=i, column=column_indices["outer_material_type5"], value="Faux Fur")
            target_sheet.cell(row=i, column=column_indices["inner_material_type1"], value=inner_material_type1_random())
            target_sheet.cell(row=i, column=column_indices["material_type1"], value=material_type1_random())
            target_sheet.cell(row=i, column=column_indices["fabric_type"], value=item["水洗唛材质"])
            target_sheet.cell(row=i, column=column_indices["fur_description"], value=item["水洗唛材质"])
            target_sheet.cell(row=i, column=column_indices["pocket_description"], value=pocket_description)
            target_sheet.cell(row=i, column=column_indices["pattern_name"], value=pattern_name_random())
            target_sheet.cell(row=i, column=column_indices["belt_style"], value=belt_style_random())
            target_sheet.cell(row=i, column=column_indices["material_type2"], value=material_type2_random())
            target_sheet.cell(row=i, column=column_indices["material_type3"], value=material_type3_random())
            target_sheet.cell(row=i, column=column_indices["control_type"], value=control_type_random())
            target_sheet.cell(row=i, column=column_indices["front_style"], value=front_style_random())
            target_sheet.cell(row=i, column=column_indices["rise_style"], value=rise_style_random())
            target_sheet.cell(row=i, column=column_indices["toe_style"], value=toe_style_random())
            target_sheet.cell(row=i, column=column_indices["lining_description"], value=item["水洗唛材质"])


            # Sell_related
            target_sheet.cell(row=i, column=column_indices["quantity"], value=199)
            target_sheet.cell(row=i, column=column_indices["fulfillment_latency"], value=2)
            target_sheet.cell(row=i, column=column_indices["merchant_shipping_group_name"], value=shipping_cost)
            target_sheet.cell(row=i, column=column_indices["item_package_quantity"], value=1)
            target_sheet.cell(row=i, column=column_indices["number_of_items"], value=1)
            target_sheet.cell(row=i, column=column_indices["condition_type"], value="New")

            # Write Brand
            target_sheet.cell(row=i, column=column_indices["brand_name"], value=brand)
            target_sheet.cell(row=i, column=column_indices["manufacturer"], value=brand)
            target_sheet.cell(row=i, column=column_indices["model_name"], value=f"{brand}{Product_Type_Map[product_Type]}{str(int(time.time() // 86400))}")

            # Write Key Word
            target_sheet.cell(row=i, column=column_indices["bullet_point1"], value=generate_Bullet(key_word_array3,key_word_array4))
            target_sheet.cell(row=i, column=column_indices["bullet_point2"], value=generate_Bullet(key_word_array3,key_word_array4))
            target_sheet.cell(row=i, column=column_indices["bullet_point3"], value=generate_Bullet(key_word_array3,key_word_array4))
            target_sheet.cell(row=i, column=column_indices["bullet_point4"], value=generate_Bullet(key_word_array3,key_word_array4))
            target_sheet.cell(row=i, column=column_indices["bullet_point5"], value=generate_Bullet(key_word_array3,key_word_array4))
            target_sheet.cell(row=i, column=column_indices["product_description"], value=generate_Bullet(key_word_array3,key_word_array4))
            target_sheet.cell(row=i, column=column_indices["generic_keywords"], value=generate_Generic(key_word_array3,key_word_array4))

            ws.cell(row=i-2, column=1, value=item['SKU'])
            ws.cell(row=i-2, column=2, value=sku)

            # When Child
            if contains_alpha:  # Child
                # subTitle
                target_sheet.cell(row=i, column=column_indices["item_name"], value=generate_Title(key_word_array1, key_word_array2, Product_Type_Map[product_Type], pattern_map[random.choice(patterns)], _color_, size, item["水洗唛材质"], title_include_fields))
                
                # Sku_Ralated
                target_sheet.cell(row=i, column=column_indices["item_sku"], value=sku)
                target_sheet.cell(row=i, column=column_indices["parent_child"], value="Child")
                target_sheet.cell(row=i, column=column_indices["parent_sku"], value=current_parent_sku)
                target_sheet.cell(row=i, column=column_indices["part_number"], value=sku[:len(prefix)+12])
                target_sheet.cell(row=i, column=column_indices["relationship_type"], value="Variation")

                # Price_Related
                target_sheet.cell(row=i, column=column_indices["standard_price"], value=price_with_margin)
                target_sheet.cell(row=i, column=column_indices["product_site_launch_date"], value=current_Time())
                target_sheet.cell(row=i, column=column_indices["merchant_release_date"], value=current_Time())
                target_sheet.cell(row=i, column=column_indices["sale_from_date"], value=current_Time())
                target_sheet.cell(row=i, column=column_indices["sale_end_date"], value=sale_End_Time())
                target_sheet.cell(row=i, column=column_indices["list_price"], value=price_with_margin+10)
                target_sheet.cell(row=i, column=column_indices["sale_price"], value=price_with_margin-1)

                # Size_Related
                target_sheet.cell(row=i, column=column_indices["size_name"], value=size)
                target_sheet.cell(row=i, column=column_indices["size_map"], value=size)
                target_sheet.cell(row=i, column=column_indices["apparel_size"], value=size)
                target_sheet.cell(row=i, column=column_indices["shirt_size"], value=size)
                target_sheet.cell(row=i, column=column_indices["bottoms_size"], value=size)
                target_sheet.cell(row=i, column=column_indices["skirt_size"], value=size)
                target_sheet.cell(row=i, column=column_indices["apparel_body_type"], value="Regular")
                target_sheet.cell(row=i, column=column_indices["apparel_height_type"], value="Regular")
                target_sheet.cell(row=i, column=column_indices["apparel_size_system"], value=site_Map[site])
                target_sheet.cell(row=i, column=column_indices["apparel_size_class"], value="Alpha")
                target_sheet.cell(row=i, column=column_indices["shirt_size_class"], value="Alpha")
                target_sheet.cell(row=i, column=column_indices["shirt_size_system"], value=site_Map[site])
                target_sheet.cell(row=i, column=column_indices["shirt_body_type"], value="Regular")
                target_sheet.cell(row=i, column=column_indices["shirt_height_type"], value="Regular")
                target_sheet.cell(row=i, column=column_indices["bottoms_size_system"], value=site_Map[site])
                target_sheet.cell(row=i, column=column_indices["bottoms_size_class"], value="Alpha")
                target_sheet.cell(row=i, column=column_indices["bottoms_body_type"], value="Regular")
                target_sheet.cell(row=i, column=column_indices["bottoms_height_type"], value="Regular")
                target_sheet.cell(row=i, column=column_indices["skirt_size_system"], value="US")
                target_sheet.cell(row=i, column=column_indices["skirt_size_class"], value="Alpha")
                target_sheet.cell(row=i, column=column_indices["skirt_body_type"], value="Regular")
                target_sheet.cell(row=i, column=column_indices["skirt_height_type"], value="Regular")
                
                # 计算当前无前缀 sku 长度
                import re
                pattern = r'([A-Za-z]+)(\d+)'
                matches = re.match(pattern, item['SKU'])
                no_prefix_sku_length = len(matches.group(1) + matches.group(2)) + 1

                color = item["颜色"]
                if color:
                    if item['SKU'][:no_prefix_sku_length] != last_sku_no_prefix or color != last_color:
                        color_prefix_map[color] = color_prefix_map.get(color, 0) + 1
                        color_prefix = f"{color_prefix_map[color]:02d}-"
                        last_color = color
                        last_sku_no_prefix = item['SKU'][:no_prefix_sku_length]

                    target_sheet.cell(row=i, column=column_indices["color_map"], value=_color_)
                    target_sheet.cell(row=i, column=column_indices["color_name"], value=color_prefix + color)
            else:  # Parent
                current_parent_sku = sku
                
                target_sheet.cell(row=i, column=column_indices["item_sku"], value=sku)
                target_sheet.cell(row=i, column=column_indices["parent_child"], value="Parent")
                target_sheet.cell(row=i, column=column_indices["part_number"], value=sku[:len(prefix)+12])
                last_color = None
                last_sku_no_prefix = None
                color_prefix_map = {}

        target_wb.save(target_file_path)
        messagebox.showinfo("成功", f"数据已成功复制并处理到目标文件: {target_file_path}")

        output_file = os.path.join(os.path.dirname(target_file_path), f"{prefix}_sku_binding.xlsx")
        wb.save(output_file)
        messagebox.showinfo("成功", f"绑定 SKU 文件已保存到: {output_file}")

    except Exception as e:
        messagebox.showerror("错误", f"处理文件时出错: {e}")
