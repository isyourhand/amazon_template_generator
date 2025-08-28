import pandas as pd
from openpyxl import load_workbook

# 加载 A.xlsx 中的 Second 表
file_a = "关键词生成工具.xlsx"
file_b = "B.xlsx"

cols = ["item_name", "bullet_point1", "bullet_point2", "bullet_point3", "bullet_point4", "bullet_point5", "generic_keywords"]

# 使用 pandas 读取关键词
df_a = pd.read_excel(file_a, sheet_name="堆词生成页", usecols=cols)

# 加载 B.xlsx 并定位到 Third 表
wb_b = load_workbook(file_b)
ws_b = wb_b["Third"]

# 将关键词从 A 的 DataFrame 写入到 B 的表格
for row_index, row_data in enumerate(df_a.values, start=2):  # 从第一行开始
    for col_index, value in enumerate(row_data, start=7):   # 从 B 列（列号 2）开始
        ws_b.cell(row=row_index, column=col_index, value=value)

# 保存 B.xlsx
wb_b.save("B.xlsx")
print("关键词已成功复制并保存到 B.xlsx")
